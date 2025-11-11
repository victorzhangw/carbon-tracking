"""
碳排放追蹤系統 - 路由
"""

from flask import Blueprint, render_template, request, jsonify
from modules.carbon_tracking.database_carbon_tracking import CarbonTrackingDB
from datetime import datetime, timedelta

carbon_bp = Blueprint('carbon', __name__, url_prefix='/carbon')
db = CarbonTrackingDB()

@carbon_bp.route('/')
def index():
    """碳排放追蹤首頁"""
    return render_template('carbon_tracking/index.html')

@carbon_bp.route('/dashboard')
def dashboard():
    """儀表板頁面"""
    return render_template('carbon_tracking/dashboard.html')

@carbon_bp.route('/visit-records')
def visit_records():
    """訪視記錄頁面"""
    return render_template('carbon_tracking/visit_records.html')

@carbon_bp.route('/add-visit')
def add_visit():
    """新增訪視記錄頁面"""
    return render_template('carbon_tracking/add_visit.html')

@carbon_bp.route('/edit-visit')
def edit_visit():
    """編輯訪視記錄頁面"""
    return render_template('carbon_tracking/edit_visit.html')

@carbon_bp.route('/statistics')
def statistics():
    """統計報表頁面"""
    return render_template('carbon_tracking/statistics.html')

@carbon_bp.route('/test-pwa')
def test_pwa():
    """PWA 功能測試頁面"""
    from flask import send_file
    return send_file('test_pwa.html')

# API 端點

@carbon_bp.route('/api/visit-records', methods=['GET'])
def get_visit_records():
    """取得訪視記錄列表（支援搜尋和篩選）"""
    # 取得參數
    limit = request.args.get('limit', 100, type=int)
    keyword = request.args.get('keyword', '').strip()
    worker_id = request.args.get('worker_id', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    transport_type = request.args.get('transport_type', '').strip()
    
    # 如果有任何篩選條件，使用篩選查詢
    if keyword or worker_id or start_date or end_date or transport_type:
        records = db.search_visit_records(
            keyword=keyword,
            worker_id=worker_id,
            start_date=start_date,
            end_date=end_date,
            transport_type=transport_type,
            limit=limit
        )
    else:
        records = db.get_all_visit_records(limit)
    
    return jsonify({'success': True, 'data': records})

@carbon_bp.route('/api/visit-records', methods=['POST'])
def create_visit_record():
    """新增訪視記錄"""
    try:
        data = request.json
        record_id = db.add_visit_record(data)
        return jsonify({'success': True, 'id': record_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/visit-records/<int:record_id>', methods=['GET'])
def get_visit_record(record_id):
    """取得單筆訪視記錄"""
    try:
        record = db.get_visit_record_by_id(record_id)
        if record:
            return jsonify({'success': True, 'data': record})
        else:
            return jsonify({'success': False, 'error': '記錄不存在'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/visit-records/<int:record_id>', methods=['PUT'])
def update_visit_record(record_id):
    """更新訪視記錄"""
    try:
        data = request.json
        success = db.update_visit_record(record_id, data)
        if success:
            return jsonify({'success': True, 'message': '更新成功'})
        else:
            return jsonify({'success': False, 'error': '記錄不存在'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/visit-records/<int:record_id>', methods=['DELETE'])
def delete_visit_record(record_id):
    """刪除訪視記錄"""
    try:
        success = db.delete_visit_record(record_id)
        if success:
            return jsonify({'success': True, 'message': '刪除成功'})
        else:
            return jsonify({'success': False, 'error': '記錄不存在'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/ai-care-records', methods=['POST'])
def create_ai_care_record():
    """新增AI關懷記錄"""
    try:
        data = request.json
        record_id = db.add_ai_care_record(data)
        return jsonify({'success': True, 'id': record_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/monthly-statistics')
def get_monthly_statistics():
    """取得月度統計"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    
    stats = db.get_monthly_statistics(year, month)
    return jsonify({'success': True, 'data': stats})

@carbon_bp.route('/api/statistics-summary')
def get_statistics_summary():
    """取得統計摘要"""
    start_date = request.args.get('start_date', '2024-06-01')
    end_date = request.args.get('end_date', '2024-09-30')
    
    stats = db.get_statistics_summary(start_date, end_date)
    return jsonify({'success': True, 'data': stats})

@carbon_bp.route('/api/period-statistics')
def get_period_statistics():
    """取得期間統計（按月）"""
    start_date = request.args.get('start_date', '2024-06-01')
    end_date = request.args.get('end_date', '2024-09-30')
    
    # 解析日期
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    
    # 按月統計
    monthly_data = []
    current = start
    
    while current <= end:
        stats = db.get_monthly_statistics(current.year, current.month)
        monthly_data.append(stats)
        
        # 移到下個月
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    
    return jsonify({'success': True, 'data': monthly_data})

@carbon_bp.route('/api/transport-distribution')
def get_transport_distribution():
    """取得交通工具分布"""
    start_date = request.args.get('start_date', '2024-06-01')
    end_date = request.args.get('end_date', '2024-09-30')
    
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
    SELECT 
        transport_type,
        COUNT(*) as count,
        SUM(distance) as total_distance,
        SUM(carbon_emission) as total_emission
    FROM visit_records
    WHERE visit_date BETWEEN ? AND ?
    GROUP BY transport_type
    ''', (start_date, end_date))
    
    columns = ['transport_type', 'count', 'total_distance', 'total_emission']
    results = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({'success': True, 'data': results})

@carbon_bp.route('/api/social-worker/<worker_id>')
def get_social_worker(worker_id):
    """根據工號查詢社工姓名"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 從訪視記錄中查詢該工號最近使用的姓名
        cursor.execute('''
        SELECT social_worker_name
        FROM visit_records
        WHERE social_worker_id = ?
        ORDER BY created_at DESC
        LIMIT 1
        ''', (worker_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return jsonify({'success': True, 'name': row[0]})
        else:
            return jsonify({'success': False, 'error': '查無此工號'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/social-workers')
def get_all_social_workers():
    """取得所有社工列表"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 取得所有不重複的社工
        cursor.execute('''
        SELECT DISTINCT social_worker_id, social_worker_name
        FROM visit_records
        ORDER BY social_worker_id
        ''')
        
        workers = []
        for row in cursor.fetchall():
            workers.append({
                'id': row[0],
                'name': row[1]
            })
        
        conn.close()
        
        return jsonify({'success': True, 'data': workers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/export/excel')
def export_excel():
    """匯出訪視記錄為 Excel"""
    try:
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        import io
        
        # 取得篩選參數
        keyword = request.args.get('keyword', '').strip()
        worker_id = request.args.get('worker_id', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        transport_type = request.args.get('transport_type', '').strip()
        
        # 查詢資料
        if keyword or worker_id or start_date or end_date or transport_type:
            records = db.search_visit_records(
                keyword=keyword,
                worker_id=worker_id,
                start_date=start_date,
                end_date=end_date,
                transport_type=transport_type,
                limit=10000
            )
        else:
            records = db.get_all_visit_records(10000)
        
        # 建立 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "訪視記錄"
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="689F38", end_color="689F38", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 標題列
        headers = ['日期', '社工編號', '社工姓名', '長者編號', '長者姓名', 
                  '訪視類型', '交通工具', '里程(km)', '行駛時間(分)', 
                  '碳排放(kg)', '出發地點', '目的地點', '備註']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 資料列
        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=record['visit_date'])
            ws.cell(row=row_idx, column=2, value=record['social_worker_id'])
            ws.cell(row=row_idx, column=3, value=record['social_worker_name'])
            ws.cell(row=row_idx, column=4, value=record['elder_id'])
            ws.cell(row=row_idx, column=5, value=record.get('elder_name', ''))
            ws.cell(row=row_idx, column=6, value=record['visit_type'])
            ws.cell(row=row_idx, column=7, value=record['transport_type'])
            ws.cell(row=row_idx, column=8, value=record['distance'])
            ws.cell(row=row_idx, column=9, value=record.get('travel_time', ''))
            ws.cell(row=row_idx, column=10, value=record.get('carbon_emission', ''))
            ws.cell(row=row_idx, column=11, value=record.get('start_location', ''))
            ws.cell(row=row_idx, column=12, value=record.get('end_location', ''))
            ws.cell(row=row_idx, column=13, value=record.get('notes', ''))
        
        # 調整欄寬
        column_widths = [12, 12, 12, 12, 12, 12, 12, 10, 12, 12, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 儲存到記憶體
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 產生檔名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'訪視記錄_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@carbon_bp.route('/api/export/csv')
def export_csv():
    """匯出訪視記錄為 CSV"""
    try:
        from flask import Response
        from datetime import datetime
        import csv
        import io
        
        # 取得篩選參數
        keyword = request.args.get('keyword', '').strip()
        worker_id = request.args.get('worker_id', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        transport_type = request.args.get('transport_type', '').strip()
        
        # 查詢資料
        if keyword or worker_id or start_date or end_date or transport_type:
            records = db.search_visit_records(
                keyword=keyword,
                worker_id=worker_id,
                start_date=start_date,
                end_date=end_date,
                transport_type=transport_type,
                limit=10000
            )
        else:
            records = db.get_all_visit_records(10000)
        
        # 建立 CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 標題列
        writer.writerow(['日期', '社工編號', '社工姓名', '長者編號', '長者姓名', 
                        '訪視類型', '交通工具', '里程(km)', '行駛時間(分)', 
                        '碳排放(kg)', '出發地點', '目的地點', '備註'])
        
        # 資料列
        for record in records:
            writer.writerow([
                record['visit_date'],
                record['social_worker_id'],
                record['social_worker_name'],
                record['elder_id'],
                record.get('elder_name', ''),
                record['visit_type'],
                record['transport_type'],
                record['distance'],
                record.get('travel_time', ''),
                record.get('carbon_emission', ''),
                record.get('start_location', ''),
                record.get('end_location', ''),
                record.get('notes', '')
            ])
        
        # 產生檔名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'訪視記錄_{timestamp}.csv'
        
        # 返回 CSV（使用 UTF-8 BOM 以支援 Excel 開啟中文）
        output_bytes = '\ufeff' + output.getvalue()
        
        return Response(
            output_bytes,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
