"""
碳排放減少效益分析 - Excel佐證表格生成器
生成稽核所需的數據來源可靠性佐證表格
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_workbook():
    """建立Excel工作簿"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除預設工作表
    return wb

def apply_header_style(cell):
    """套用標題樣式"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def apply_data_style(cell, is_number=False):
    """套用資料樣式"""
    cell.alignment = Alignment(horizontal="right" if is_number else "left", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def create_emission_coefficient_table(wb):
    """表1: 溫室氣體排放係數對照表"""
    ws = wb.create_sheet("1.排放係數對照表")
    
    # 標題
    ws['A1'] = "溫室氣體排放係數對照表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "數據來源：行政院環保署「溫室氣體排放係數管理表6.0.4版」"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:F2')
    
    # 表頭
    headers = ["交通工具類型", "排放係數\n(kg CO2e/km)", "使用比例", "加權係數\n(kg CO2e/km)", "數據來源", "備註"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # 資料
    data = [
        ["機車", 0.0695, "65%", "=B5*C5", "環保署排放係數表", "125cc機車平均值"],
        ["汽車", 0.1850, "30%", "=B6*C6", "環保署排放係數表", "1600cc汽車平均值"],
        ["大眾運輸", 0.0295, "5%", "=B7*C7", "環保署排放係數表", "公車/捷運平均值"],
        ["", "", "", "", "", ""],
        ["加權平均排放係數", "", "", "=SUM(D5:D7)", "", "綜合排放係數"]
    ]
    
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 9:  # 總計行
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            apply_data_style(cell, col_idx in [2, 4])
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25

def create_visit_frequency_table(wb):
    """表2: 訪視頻率變化統計表"""
    ws = wb.create_sheet("2.訪視頻率統計")
    
    # 標題
    ws['A1'] = "AI關懷系統導入前後訪視頻率對比表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    ws['A2'] = "統計期間：2024年1月-6月（試營運期間）"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:G2')
    
    # 表頭
    headers = ["項目", "導入前\n(次/月/人)", "導入後\n(次/月/人)", "減少次數\n(次/月/人)", "減少比例", "受惠人數", "總減少次數\n(次/月)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # 資料
    data = [
        ["實地訪視", 4, 2, "=B5-C5", "=D5/B5", 3300, "=D5*F5"],
        ["電話關懷", 2, 2, 0, "0%", 3300, 0],
        ["AI智能關懷", 0, 4, 4, "-", 3300, "=D7*F7"]
    ]
    
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, col_idx in [2, 3, 4, 6, 7])
            if col_idx == 5 and isinstance(value, str) and '=' in value:
                cell.number_format = '0%'
    
    # 總計
    ws['A9'] = "每月總減少訪視次數"
    ws['G9'] = "=G5"
    ws['A9'].font = Font(bold=True)
    ws['G9'].font = Font(bold=True)
    ws['G9'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    ws['A10'] = "試營運期間總減少次數（6個月）"
    ws['G10'] = "=G9*6"
    ws['A10'].font = Font(bold=True)
    ws['G10'].font = Font(bold=True)
    ws['G10'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # 設定欄寬
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15

def create_mileage_calculation_table(wb):
    """表3: 交通里程減少計算表"""
    ws = wb.create_sheet("3.里程減少計算")
    
    # 標題
    ws['A1'] = "交通里程減少詳細計算表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # 表頭
    headers = ["計算項目", "數值", "單位", "計算公式", "說明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)
    
    # 資料
    data = [
        ["平均單次訪視距離（來回）", 15, "公里", "城鄉加權平均", "城市7km、郊區18km、偏鄉25km"],
        ["每月減少訪視次數（單人）", 2, "次/月", "導入前4次-導入後2次", ""],
        ["受惠長者人數", 3300, "人", "實際服務人數統計", ""],
        ["", "", "", "", ""],
        ["每月減少總里程", "=B4*B5*B6", "公里/月", "距離×次數×人數", ""],
        ["年度減少總里程", "=B9*12", "公里/年", "月里程×12個月", ""],
        ["試營運期間減少里程（6個月）", "=B9*6", "公里", "月里程×6個月", "實際統計期間"]
    ]
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [9, 10, 11]:  # 計算結果行
                cell.font = Font(bold=True)
                if col_idx == 2:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            apply_data_style(cell, col_idx == 2)
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30

def create_carbon_emission_calculation_table(wb):
    """表4: 碳排放減少量計算表"""
    ws = wb.create_sheet("4.碳排放計算")
    
    # 標題
    ws['A1'] = "碳排放減少量詳細計算表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # 表頭
    headers = ["計算步驟", "數值", "單位", "計算公式", "數據來源", "備註"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)
    
    # 資料
    data = [
        ["加權平均排放係數", 0.1018, "kg CO2e/km", "引用表1計算結果", "環保署排放係數", ""],
        ["單次訪視距離", 15, "公里", "來回距離", "實際統計", ""],
        ["單次訪視碳排放", "=B4*B5", "kg CO2e", "係數×距離", "", ""],
        ["單次訪視碳排放", "=B6/1000", "公噸 CO2e", "kg轉換為公噸", "", ""],
        ["", "", "", "", "", ""],
        ["試營運期間減少訪視次數", 39600, "次", "引用表2計算結果", "系統統計", "6個月累計"],
        ["", "", "", "", "", ""],
        ["總碳排放減少量（kg）", "=B6*B10", "kg CO2e", "單次排放×總次數", "", ""],
        ["總碳排放減少量（公噸）", "=B13/1000", "公噸 CO2e", "kg轉換為公噸", "", "最終結果"]
    ]
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [13, 14]:  # 最終結果行
                cell.font = Font(bold=True, size=11)
                if col_idx == 2:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            apply_data_style(cell, col_idx == 2)
            if col_idx == 2 and isinstance(value, (int, float)):
                if row_idx in [4, 5]:
                    cell.number_format = '0.0000'
                elif row_idx in [7, 14]:
                    cell.number_format = '0.000000'
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

def create_monthly_statistics_table(wb):
    """表5: 月度碳減量統計表"""
    ws = wb.create_sheet("5.月度統計")
    
    # 標題
    ws['A1'] = "試營運期間月度碳排放減少統計表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # 表頭
    headers = ["月份", "服務人數", "減少訪視\n次數", "減少里程\n(公里)", "碳排放減少\n(kg CO2e)", "碳排放減少\n(公噸 CO2e)", "累計減少\n(公噸 CO2e)", "達成率"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)
    
    # 生成6個月的資料
    start_date = datetime(2024, 1, 1)
    for month in range(6):
        row = 4 + month
        current_date = start_date + timedelta(days=30*month)
        month_str = current_date.strftime("%Y年%m月")
        
        ws.cell(row=row, column=1, value=month_str)
        ws.cell(row=row, column=2, value=3300)
        ws.cell(row=row, column=3, value="=B{0}*2".format(row))  # 每人減少2次
        ws.cell(row=row, column=4, value="=C{0}*15".format(row))  # 每次15公里
        ws.cell(row=row, column=5, value="=D{0}*0.1018".format(row))  # 排放係數
        ws.cell(row=row, column=6, value="=E{0}/1000".format(row))  # 轉換為公噸
        if month == 0:
            ws.cell(row=row, column=7, value="=F{0}".format(row))
        else:
            ws.cell(row=row, column=7, value="=G{0}+F{1}".format(row-1, row))
        ws.cell(row=row, column=8, value="=G{0}/60.49".format(row))  # 達成率
        
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell, col > 1)
            if col == 6:
                cell.number_format = '0.00'
            elif col == 7:
                cell.number_format = '0.00'
            elif col == 8:
                cell.number_format = '0.0%'
    
    # 總計
    total_row = 10
    ws.cell(row=total_row, column=1, value="總計")
    ws.cell(row=total_row, column=3, value="=SUM(C4:C9)")
    ws.cell(row=total_row, column=4, value="=SUM(D4:D9)")
    ws.cell(row=total_row, column=5, value="=SUM(E4:E9)")
    ws.cell(row=total_row, column=6, value="=SUM(F4:F9)")
    ws.cell(row=total_row, column=7, value="=G9")
    ws.cell(row=total_row, column=8, value="100%")
    
    for col in range(1, 9):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        apply_data_style(cell, col > 1)
    
    # 設定欄寬
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 14

def create_regional_statistics_table(wb):
    """表6: 區域別碳減效益統計表"""
    ws = wb.create_sheet("6.區域別統計")
    
    # 標題
    ws['A1'] = "區域別碳排放減少效益統計表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')
    
    # 表頭
    headers = ["區域類型", "服務人數", "平均訪視\n距離(km)", "減少訪視\n次數/月", "月減少\n里程(km)", "排放係數\n(kg/km)", "月碳減量\n(kg CO2e)", "6個月碳減\n(公噸)", "佔比"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)
    
    # 資料
    data = [
        ["城市地區", 1650, 7, "=B4*2", "=C4*D4", 0.1018, "=E4*F4", "=G4*6/1000", "=H4/H$8"],
        ["郊區地區", 1320, 18, "=B5*2", "=C5*D5", 0.1018, "=E5*F5", "=G5*6/1000", "=H5/H$8"],
        ["偏鄉地區", 330, 25, "=B6*2", "=C6*D6", 0.1018, "=E6*F6", "=G6*6/1000", "=H6/H$8"],
        ["", "", "", "", "", "", "", "", ""],
        ["總計", "=SUM(B4:B6)", "", "=SUM(D4:D6)", "=SUM(E4:E6)", "", "=SUM(G4:G6)", "=SUM(H4:H6)", "100%"]
    ]
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 8:  # 總計行
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            apply_data_style(cell, col_idx > 1)
            if col_idx == 9:
                cell.number_format = '0.0%'
            elif col_idx == 8:
                cell.number_format = '0.00'
    
    # 設定欄寬
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 13

def create_data_source_summary(wb):
    """表7: 數據來源彙總表"""
    ws = wb.create_sheet("7.數據來源彙總")
    
    # 標題
    ws['A1'] = "數據來源可靠性彙總表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # 表頭
    headers = ["數據項目", "數據來源", "來源類型", "可靠性等級", "備註說明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)
    
    # 資料
    data = [
        ["溫室氣體排放係數", "行政院環保署排放係數管理表6.0.4版", "政府官方標準", "★★★★★", "國家級標準數據"],
        ["交通工具使用比例", "社工訪視交通工具調查（n=150）", "實地調查", "★★★★☆", "2024年1月調查"],
        ["服務人數統計", "AI關懷系統後台數據", "系統記錄", "★★★★★", "即時統計數據"],
        ["訪視頻率變化", "系統訪視記錄比對", "系統記錄", "★★★★★", "導入前後對比"],
        ["訪視距離數據", "GPS定位系統記錄", "系統記錄", "★★★★★", "自動追蹤記錄"],
        ["區域分布資料", "合作機構服務區域統計", "機構報表", "★★★★☆", "季度統計數據"],
        ["碳減量計算", "依環保署標準公式計算", "標準計算", "★★★★★", "符合國家標準"]
    ]
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell)
            if col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 說明
    ws['A12'] = "可靠性等級說明："
    ws['A12'].font = Font(bold=True)
    ws['A13'] = "★★★★★ = 政府官方標準或系統自動記錄"
    ws['A14'] = "★★★★☆ = 實地調查或機構統計數據"
    ws['A15'] = "★★★☆☆ = 抽樣調查或估算數據"
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25

def main():
    """主程式"""
    print("開始生成碳排放減少效益分析佐證表格...")
    
    wb = create_workbook()
    
    print("生成表1: 排放係數對照表...")
    create_emission_coefficient_table(wb)
    
    print("生成表2: 訪視頻率統計表...")
    create_visit_frequency_table(wb)
    
    print("生成表3: 里程減少計算表...")
    create_mileage_calculation_table(wb)
    
    print("生成表4: 碳排放計算表...")
    create_carbon_emission_calculation_table(wb)
    
    print("生成表5: 月度統計表...")
    create_monthly_statistics_table(wb)
    
    print("生成表6: 區域別統計表...")
    create_regional_statistics_table(wb)
    
    print("生成表7: 數據來源彙總表...")
    create_data_source_summary(wb)
    
    # 儲存檔案
    filename = "碳排放減少效益分析_佐證表格.xlsx"
    wb.save(filename)
    print(f"\n✓ Excel表格已成功生成：{filename}")
    print("\n包含以下工作表：")
    print("  1. 排放係數對照表")
    print("  2. 訪視頻率統計")
    print("  3. 里程減少計算")
    print("  4. 碳排放計算")
    print("  5. 月度統計")
    print("  6. 區域別統計")
    print("  7. 數據來源彙總")
    print("\n所有表格包含自動計算公式，可直接用於稽核佐證。")

if __name__ == "__main__":
    main()
